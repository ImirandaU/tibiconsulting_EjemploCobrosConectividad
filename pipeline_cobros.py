#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Pipeline de Cobros por Conectividad IoT — repositorio de referencia
=====================================================================
Caso de arquitectura (anonimizado) construido por TIBI Consulting.

CONTEXTO
--------
Una empresa de conectividad celular para dispositivos IoT factura
mensualmente según la cantidad de SIMs activas por plan, repartidas en
múltiples países y operadoras, cada una con su propia moneda. Antes de
este pipeline, el cálculo lo armaban a mano y por separado dos áreas
distintas, sin una fuente de verdad común ni trazabilidad.

Este repositorio muestra el pipeline que reemplazó ese proceso manual:
tres capas de datos, auditables de punta a punta, con la lógica de
negocio que cambia seguido (precios, operadoras, reglas de clasificación)
viviendo en base de datos en vez de en el código.

    archivo mensual (Excel)
        └──▶ capa RAW tipada (landing)
                 └──▶ cruce con el maestro de precios (mantenedor)
                          └──▶ resultado de cobros + Excel ejecutivo

PRINCIPIO DE DISEÑO CENTRAL
----------------------------
Agregar un país, una operadora o un plan nuevo es un INSERT en la tabla
`maestro_precios`, no un cambio de código ni un despliegue. El mantenedor
lo opera el área de negocio; el código no vuelve a tocarse.

NOTA IMPORTANTE
----------------
Este es un repositorio de REFERENCIA con fines de portafolio. Los nombres
de esquema, las rutas de archivos y los datos de ejemplo (sql/04_seed_ejemplo.sql)
son ILUSTRATIVOS: no corresponden a ningún cliente, tarifa o acuerdo comercial
real. La arquitectura, los patrones y la lógica de procesamiento sí son fieles
al sistema real en producción.

USO
---
    python pipeline_cobros.py        -> abre el menú interactivo

Configuración vía variables de entorno (ver .env.example):
    DB_HOST, DB_PORT, DB_NAME, DB_USER, DB_PASSWORD
    INPUT_DIR   -> carpeta con los Excel mensuales
    OUTPUT_DIR  -> carpeta donde se dejan los reportes generados
"""

import os
import sys
import re
import glob
import calendar
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
from sqlalchemy import create_engine, text
from dotenv import load_dotenv

load_dotenv()  # carga variables desde un archivo .env local, si existe


# ============================================================================
# CONFIGURACIÓN  (100% por variables de entorno — sin rutas ni credenciales
# hardcodeadas, para que el repositorio sea seguro de publicar y fácil de
# adaptar a cualquier entorno)
# ============================================================================
INPUT_DIR  = os.getenv("INPUT_DIR", "./data/input")
OUTPUT_DIR = os.getenv("OUTPUT_DIR", "./data/output")

# Esquemas de base de datos (nombres configurables; estos son los default).
RAW_SCHEMA = os.getenv("RAW_SCHEMA", "raw")           # capa de landing
DB_SCHEMA  = os.getenv("DB_SCHEMA", "billing")         # capa de reportería

TABLA_MAESTRO   = "maestro_precios"
TABLA_RESULTADO = "resultado_cobros"
TABLA_LOG       = "log_ejecucion"
RAW_TABLA       = "cuentas_raw"

LOG_SCHEMA = DB_SCHEMA   # la bitácora vive junto al resultado

# Hoja de Excel que contiene los datos (las demás se ignoran; ver
# leer_hoja_datos: algunos meses traían hojas vacías puestas antes de la real).
HOJA_DATOS = "Worksheet"

# Valores de "cuenta padre" que se excluyen del conteo (cuentas internas,
# de administración, o sin asignar). Ilustrativo: ajustar a tu propio dataset.
PAIS_PADRE_EXCLUIR = ["Cuentas Internas (0000)", "Admin (0)", "()"]

MESES = {
    "enero": "01", "febrero": "02", "marzo": "03", "abril": "04",
    "mayo": "05", "junio": "06", "julio": "07", "agosto": "08",
    "septiembre": "09", "setiembre": "09", "octubre": "10",
    "noviembre": "11", "diciembre": "12",
}
NUM_A_MES = {v: k for k, v in MESES.items() if k != "setiembre"}

# Columnas esperadas en el Excel mensual -> columna SQL de la capa RAW.
# tipos: text | int | bigint | numeric  -> coerción segura (dato sucio -> NULL)
COLUMNAS_RAW = [
    ("Ciclo Facturacion",        "ciclo_facturacion",        "text"),
    ("Cuenta Padre",             "cuenta_padre",             "text"),
    ("CompanyId",                "company_id",               "bigint"),
    ("Empresa",                  "empresa",                  "text"),
    ("Tipo Sim",                 "tipo_sim",                 "text"),
    ("Plan",                     "plan",                     "text"),
    ("Código Plan",              "codigo_plan",              "text"),
    ("Total Sims",               "total_sims",               "int"),
    ("Precio Plan",              "precio_plan",              "numeric"),
    ("Pool Disponible",          "pool_disponible",          "numeric"),
    ("Total Consumo MB",         "total_consumo_mb",         "numeric"),
    ("Total Sobreconsumo",       "total_sobreconsumo",       "numeric"),
    ("Total Cobro Plan",         "total_cobro_plan",         "numeric"),
    ("Total Cobro Sobreconsumo", "total_cobro_sobreconsumo", "numeric"),
    ("Total Cobro SMS",          "total_cobro_sms",          "numeric"),
    ("Total Cobro Voz",          "total_cobro_voz",          "numeric"),
    ("Total Cobro",              "total_cobro",              "numeric"),
    ("Moneda",                   "moneda",                   "text"),
    ("Comentarios",              "comentarios",               "text"),
]

# Regex de apoyo para leer el tamaño de datos de un plan y evaluar reglas
# de clasificación (ver _cumple_regla).
_RE_TAMANO = re.compile(r'(\d+(?:[.,]\d+)?)\s*(gb|mb)', re.IGNORECASE)
_RE_REGLA  = re.compile(r'(<=|>=|<|>)\s*(\d+(?:[.,]\d+)?)\s*(gb|mb)', re.IGNORECASE)
_MONEDA_INVALIDA = {"", "SIN_DATO", "SINDATO", "NONE", "NAN"}


# ============================================================================
# BASE DE DATOS
# ============================================================================
def conectar_db():
    """Crea la conexión a PostgreSQL a partir de variables de entorno."""
    try:
        host = os.environ["DB_HOST"]
        port = os.getenv("DB_PORT", "5432")
        name = os.environ["DB_NAME"]
        user = os.environ["DB_USER"]
        pwd  = os.environ["DB_PASSWORD"]
    except KeyError as e:
        print(f"❌ Falta la variable de entorno {e}. Revisa tu archivo .env "
              f"(ver .env.example).")
        return None

    try:
        engine = create_engine(f"postgresql+psycopg2://{user}:{pwd}@{host}:{port}/{name}")
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        print(f"✅ Conexión a PostgreSQL exitosa ({host}:{port}/{name})")
        return engine
    except Exception as e:
        print(f"❌ Error conectando a PostgreSQL: {e}")
        return None


def _ref(tabla: str, schema) -> str:
    """Nombre de tabla calificado por schema."""
    return f'"{schema}".{tabla}' if schema else tabla


def _eliminar_periodo(engine, tabla: str, periodo: str, schema) -> bool:
    """DELETE por periodo + reset de secuencia. Permite reprocesar sin duplicar."""
    ref = _ref(tabla, schema)
    try:
        with engine.begin() as conn:
            result = conn.execute(text(f'DELETE FROM {ref} WHERE periodo = :periodo'),
                                  {"periodo": periodo})
            filas = result.rowcount or 0
            if filas > 0:
                print(f"  🗑️ Eliminados {filas} registros previos del periodo {periodo}")
            if engine.dialect.name == "postgresql":
                conn.execute(text(f"""
                    SELECT setval(pg_get_serial_sequence('{ref}', 'id'),
                                  COALESCE((SELECT MAX(id) FROM {ref}), 0) + 1, false)
                """))
        return True
    except Exception as e:
        print(f"  ❌ Error eliminando periodo en {ref}: {e}")
        return False


_DEFAULT_SCHEMA = object()  # sentinela: distingue "no pasado" de None (schema public)


def _log_ejecucion(conn, periodo, archivo, filas, sims, modo, schema=_DEFAULT_SCHEMA):
    """Registra cada corrida (carga RAW o cálculo de reporte) en la bitácora."""
    if schema is _DEFAULT_SCHEMA:
        schema = LOG_SCHEMA
    conn.execute(
        text(f"""INSERT INTO {_ref(TABLA_LOG, schema)}
                 (fecha_ejecucion, periodo, archivo_origen,
                  filas_generadas, sims_totales, modo)
                 VALUES (:fe, :p, :arch, :filas, :sims, :modo)"""),
        {"fe": datetime.now(), "p": periodo,
         "arch": (Path(archivo).name if archivo else None),
         "filas": int(filas), "sims": int(sims), "modo": modo},
    )


# ============================================================================
# LECTURA DE EXCEL  (blindada contra hojas vacías y archivos de bloqueo)
# ============================================================================
def _sin_tildes(texto: str) -> str:
    norm = unicodedata.normalize("NFKD", str(texto))
    return "".join(c for c in norm if not unicodedata.combining(c)).lower()


def leer_hoja_datos(path_excel: str) -> pd.DataFrame:
    """
    Lee la hoja real de datos. Usa 'Worksheet' si existe; si no, la hoja con
    más columnas. Blindaje real: algunos meses traían una hoja vacía puesta
    ANTES de la hoja de datos real, y una lectura ingenua devolvía cero filas.
    """
    xl = pd.ExcelFile(path_excel)
    if HOJA_DATOS in xl.sheet_names:
        return xl.parse(HOJA_DATOS)
    mejor_nombre, mejor_df = None, None
    for sn in xl.sheet_names:
        d = xl.parse(sn)
        if mejor_df is None or d.shape[1] > mejor_df.shape[1]:
            mejor_nombre, mejor_df = sn, d
    print(f"  ⚠️ '{Path(path_excel).name}' sin hoja '{HOJA_DATOS}'; uso '{mejor_nombre}'")
    return mejor_df if mejor_df is not None else pd.DataFrame()


# ============================================================================
# CAPA RAW  (landing tipada)
# ============================================================================
def normalizar_raw(df: pd.DataFrame, periodo: str, archivo: str) -> pd.DataFrame:
    """Renombra a snake_case, tipa con coerción segura y agrega trazabilidad."""
    out = pd.DataFrame(index=df.index)
    for exc, sql, tipo in COLUMNAS_RAW:
        serie = df[exc] if exc in df.columns else pd.Series([pd.NA] * len(df), index=df.index)
        if tipo in ("int", "bigint"):
            out[sql] = pd.to_numeric(serie, errors="coerce").round().astype("Int64")
        elif tipo == "numeric":
            out[sql] = pd.to_numeric(serie, errors="coerce")
        else:
            out[sql] = serie.astype("string")
    out["periodo"] = periodo
    out["archivo_origen"] = Path(archivo).name
    out["fecha_carga"] = datetime.now()
    return out


def cargar_raw_db(engine, df_raw: pd.DataFrame, periodo: str, archivo: str,
                  schema=RAW_SCHEMA) -> int:
    """Inserta la RAW del periodo (NA -> None) y registra la bitácora."""
    if df_raw.empty:
        print("  ⚠️ Sin filas que cargar")
        return 0
    df_ins = df_raw.astype(object).where(pd.notnull(df_raw), None)
    sims = int(pd.to_numeric(df_raw["total_sims"], errors="coerce").fillna(0).sum())
    try:
        with engine.begin() as conn:
            df_ins.to_sql(RAW_TABLA, conn, schema=schema, if_exists="append", index=False)
            _log_ejecucion(conn, periodo, archivo, len(df_ins), sims, "raw")
        print(f"  ✅ RAW cargada: {len(df_ins):,} filas  (SIMs={sims:,})")
        return len(df_ins)
    except Exception as e:
        print(f"  ❌ Error cargando RAW: {e}")
        return 0


def procesar_raw(engine, archivo: str, periodo=None, commit=True, schema=RAW_SCHEMA) -> int:
    """Lee un Excel mensual y lo sube tipado a la capa RAW."""
    periodo = periodo or derivar_periodo(archivo)
    print(f"\n📥 RAW periodo {periodo}  ←  {Path(archivo).name}")
    df = leer_hoja_datos(archivo)
    if df.empty:
        print("  ⚠️ El archivo no tiene datos legibles")
        return 0
    df_raw = normalizar_raw(df, periodo, archivo)
    print(f"  📊 Filas leídas: {len(df_raw):,}  |  columnas: {len(COLUMNAS_RAW)}")
    if not commit:
        print("  ℹ️ DRY-RUN: no se escribió la RAW")
        return 0
    if _eliminar_periodo(engine, RAW_TABLA, periodo, schema):
        return cargar_raw_db(engine, df_raw, periodo, archivo, schema)
    return 0


# ============================================================================
# PERIODO Y ARCHIVOS
# ============================================================================
def derivar_periodo(path_excel: str) -> str:
    """
    Periodo YYYYMM desde el NOMBRE del archivo (año 20XX + mes en texto).
    A propósito NO usa un prefijo numérico del nombre: en la práctica ese
    prefijo se escribe a mano y puede venir mal (ej. mes equivocado).
    Ej: 'Cuentas_2026_Mayo.xlsx' -> '202605'
    """
    nombre = _sin_tildes(Path(path_excel).stem)
    m_anio = re.search(r"(20\d{2})", nombre)
    if not m_anio:
        raise ValueError(f"No pude detectar el año (20XX) en: {path_excel}")
    anio = m_anio.group(1)
    mes = next((num for nom, num in MESES.items() if nom in nombre), None)
    if mes is None:
        raise ValueError(f"No pude detectar el mes (texto) en: {path_excel}")
    return f"{anio}{mes}"


def fecha_corte(periodo: str) -> datetime:
    anio, mes = int(periodo[:4]), int(periodo[4:])
    return datetime(anio, mes, calendar.monthrange(anio, mes)[1])


def _listar_cuentas(carpeta: str):
    """
    Lista los Excel mensuales, ignorando los archivos temporales de bloqueo
    que Office crea cuando un archivo está abierto (prefijo '~$...').
    """
    encontrados = []
    for f in glob.glob(os.path.join(carpeta, "*.xlsx")):
        nombre = os.path.basename(f)
        if nombre.startswith("~$"):
            continue
        if "cuentas" in _sin_tildes(nombre):
            encontrados.append(f)
    return encontrados


def localizar_archivo_periodo(periodo: str) -> str:
    """Ubica el Excel mensual del periodo por AÑO + MES (nombre)."""
    anio, mes = periodo[:4], periodo[4:]
    mes_nom = NUM_A_MES.get(mes)
    cands = [f for f in _listar_cuentas(INPUT_DIR)
             if anio in os.path.basename(f) and mes_nom in _sin_tildes(os.path.basename(f))]
    if not cands:
        disponibles = "\n   ".join(os.path.basename(f) for f in _listar_cuentas(INPUT_DIR)) or "(ninguno)"
        raise FileNotFoundError(f"No encontré archivo para {periodo} en {INPUT_DIR}\n"
                                f"   Disponibles:\n   {disponibles}")
    if len(cands) > 1:
        raise FileNotFoundError(f"Hay más de un archivo para {periodo}: {cands}")
    return cands[0]


def periodos_en_archivos():
    """Lista (periodo, ruta) de todos los archivos disponibles, ordenado."""
    pares = []
    for f in _listar_cuentas(INPUT_DIR):
        try:
            pares.append((derivar_periodo(f), f))
        except ValueError:
            continue
    return sorted(pares)


def periodos_en_raw(engine, schema=RAW_SCHEMA):
    """Periodos distintos presentes en la capa RAW, ordenados."""
    q = text(f"SELECT DISTINCT periodo FROM {_ref(RAW_TABLA, schema)} ORDER BY periodo")
    return [str(r[0]).strip() for r in pd.read_sql(q, engine).itertuples(index=False, name=None)]


# ============================================================================
# CRUCE: maestro de precios x capa RAW -> resultado
# ============================================================================
def _parse_tamano_gb(nombre: str):
    """Extrae el tamaño de datos del nombre del plan, en GB. None si no hay."""
    ms = _RE_TAMANO.findall(str(nombre))
    if not ms:
        return None
    num, unit = ms[-1]
    val = float(num.replace(",", "."))
    return val if unit.lower() == "gb" else val / 1024.0


def _cumple_regla(plan: str, regla) -> bool:
    """
    Evalúa la regla de clasificación por tamaño (ej. '< 1GB', '>= 1GB').
    Sin regla -> True (solo aplica el prefijo). Sin tamaño parseable -> False.

    Caso real que motivó esto: una operadora ofrece planes que comparten el
    MISMO nombre/prefijo pero deben clasificarse en dos categorías de precio
    distintas según el volumen de datos contratado. En vez de listar cada
    plan a mano, se define la regla una vez en el maestro de precios.
    """
    if regla is None or (isinstance(regla, float) and pd.isna(regla)) or not str(regla).strip():
        return True
    m = _RE_REGLA.search(str(regla))
    if not m:
        return True
    op, num, unit = m.groups()
    umbral = float(num.replace(",", "."))
    if unit.lower() == "mb":
        umbral /= 1024.0
    g = _parse_tamano_gb(plan)
    if g is None:
        return False
    return {"<": g < umbral, "<=": g <= umbral, ">": g > umbral, ">=": g >= umbral}[op]


def precios_vigentes(engine, corte: datetime, schema=DB_SCHEMA) -> pd.DataFrame:
    """
    Fila vigente de cada llave (país, operadora, tipo) a la fecha de corte:
    la de fecha más reciente que sea <= corte. El maestro es historizado
    (append-only): cada cambio de precio se agrega como fila nueva con su
    propia fecha de vigencia, nunca se sobrescribe. Así, si se reprocesa un
    mes antiguo, se usa el precio que estaba vigente en ese momento.
    """
    maestro = pd.read_sql(f"SELECT * FROM {_ref(TABLA_MAESTRO, schema)}", engine)
    if maestro.empty:
        return maestro
    maestro["fecha"] = pd.to_datetime(maestro["fecha"], errors="coerce")
    maestro["fecha_carga"] = pd.to_datetime(maestro["fecha_carga"], format="mixed",
                                            utc=True, errors="coerce")
    vig = maestro[maestro["fecha"] <= pd.Timestamp(corte)].copy()
    vig = (vig.sort_values(["fecha", "fecha_carga"])
              .groupby(["pais", "operador", "tipo_operador"], as_index=False)
              .tail(1))
    return vig


def leer_raw_planes(engine, periodo: str, schema_raw=RAW_SCHEMA):
    """Lee plan + total_sims de la RAW del periodo, filtrando cuentas internas."""
    ref = _ref(RAW_TABLA, schema_raw)
    df = pd.read_sql(
        text(f"SELECT plan, total_sims, cuenta_padre, archivo_origen "
             f"FROM {ref} WHERE periodo = :p"),
        engine, params={"p": periodo})
    if df.empty:
        return df, None
    archivo = df["archivo_origen"].dropna().iloc[0] if df["archivo_origen"].notna().any() else f"(raw {periodo})"
    df = df[~df["cuenta_padre"].isin(PAIS_PADRE_EXCLUIR)].copy()
    df["total_sims"] = pd.to_numeric(df["total_sims"], errors="coerce").fillna(0)
    return df, archivo


def _sims_de(df_planes: pd.DataFrame, prefijo, regla) -> int:
    """Suma total_sims de los planes que matchean el prefijo (y la regla, si hay)."""
    if not prefijo:
        return 0
    sub = df_planes[df_planes["plan"].str.startswith(prefijo, na=False)]
    if sub.empty:
        return 0
    if regla is not None and not (isinstance(regla, float) and pd.isna(regla)) and str(regla).strip():
        sub = sub[sub["plan"].map(lambda p: _cumple_regla(p, regla))]
    return int(sub["total_sims"].sum())


def construir_resultado(precios: pd.DataFrame, df_planes: pd.DataFrame,
                        periodo: str, fecha_ejecucion: datetime) -> pd.DataFrame:
    """
    Recorre cada fila VIGENTE del maestro de precios y arma el resultado:
    suma SIMs por prefijo (+regla) y calcula monto = precio * sims.

    Esto es lo que hace que agregar un país/operadora nuevo sea un simple
    INSERT: el código no conoce ninguna lista fija de operadoras, solo
    recorre lo que encuentra en la tabla.
    """
    filas = []
    for _, r in precios.iterrows():
        prefijo = r.get("plan_prefijo")
        prefijo = None if (prefijo is None or (isinstance(prefijo, float) and pd.isna(prefijo))) else str(prefijo)
        regla = r.get("regla")
        sims = _sims_de(df_planes, prefijo, regla)

        precio = r.get("precio_unitario")
        precio_valido = pd.notna(precio)
        moneda_raw = r.get("moneda")
        moneda_valida = (pd.notna(moneda_raw)
                         and str(moneda_raw).strip().upper() not in _MONEDA_INVALIDA)

        fila = {
            "fecha_ejecucion": fecha_ejecucion, "periodo": periodo,
            "pais": r.get("pais"), "operador": r.get("operador"),
            "tipo_operador": r.get("tipo_operador"), "plan_prefijo": prefijo,
            "sims_total": sims, "precio_unitario": None, "moneda": None,
            "monto": None, "comentario": None,
        }
        if not prefijo:
            fila["comentario"] = "Sin plan_prefijo en el maestro (no se puede mapear)"
        elif precio_valido and moneda_valida:
            precio = float(precio)
            fila["precio_unitario"] = precio
            fila["moneda"] = str(moneda_raw).strip()
            fila["monto"] = round(precio * sims, 4)
        else:
            fila["precio_unitario"] = float(precio) if precio_valido else None
            fila["comentario"] = "Sin precio/moneda válidos en el maestro (solo SIMs)"
        filas.append(fila)

    columnas = ["fecha_ejecucion", "periodo", "pais", "operador", "tipo_operador",
                "plan_prefijo", "sims_total", "precio_unitario", "moneda",
                "monto", "comentario"]
    return pd.DataFrame(filas, columns=columnas)


def cargar_resultado_db(engine, resultado: pd.DataFrame, periodo: str,
                        archivo_origen: str, schema=DB_SCHEMA) -> int:
    if resultado.empty:
        return 0
    try:
        with engine.begin() as conn:
            resultado.to_sql(TABLA_RESULTADO, conn, schema=schema, if_exists="append", index=False)
            _log_ejecucion(conn, periodo, archivo_origen, len(resultado),
                           int(resultado["sims_total"].sum()), "reporte", schema)
        print(f"  ✅ Resultado cargado: {len(resultado)} registros")
        return len(resultado)
    except Exception as e:
        print(f"  ❌ Error cargando resultado: {e}")
        return 0


def procesar_reporte(engine, periodo: str, commit=True, salida=None,
                     schema=DB_SCHEMA, schema_raw=RAW_SCHEMA):
    """Cruza la RAW del periodo con el maestro y genera resultado + Excel."""
    fecha_ejecucion = datetime.now()
    corte = fecha_corte(periodo)
    print(f"\n🧮 REPORTE periodo {periodo}  (corte {corte.date()})")

    df_planes, archivo_origen = leer_raw_planes(engine, periodo, schema_raw)
    if df_planes is None or df_planes.empty:
        print(f"  ⚠️ No hay data RAW para {periodo}. Carga primero la RAW (opción 1).")
        return None

    precios = precios_vigentes(engine, corte, schema)
    if precios.empty:
        print("  ⚠️ El maestro de precios está vacío. No hay nada que cruzar.")
        return None
    resultado = construir_resultado(precios, df_planes, periodo, fecha_ejecucion)
    print(f"  📋 {len(resultado)} llaves del maestro cruzadas")

    if salida is None:
        salida = os.path.join(OUTPUT_DIR, f"reporte_cobros_{periodo}.xlsx")
    generar_excel(resultado, periodo, archivo_origen, salida)
    print(f"  💾 Excel: {salida}")

    if commit:
        if _eliminar_periodo(engine, TABLA_RESULTADO, periodo, schema):
            cargar_resultado_db(engine, resultado, periodo, archivo_origen, schema)
    else:
        print("  ℹ️ DRY-RUN: no se escribió el resultado en la BD")
    return resultado


# ============================================================================
# REPORTE EXCEL
# ============================================================================
def generar_excel(resultado: pd.DataFrame, periodo: str, archivo_origen: str, path_salida: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    azul = PatternFill("solid", fgColor="1F4E78")
    gris = PatternFill("solid", fgColor="F2F2F2")
    blanco_bold = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)
    normal = Font()
    centro = Alignment(horizontal="center")
    fmt_num, fmt_dec = "#,##0", "#,##0.00"

    wb = Workbook()
    ws = wb.active
    ws.title = "Detalle"
    encabezados = ["Periodo", "Fecha Ejecución", "País", "Operador", "Tipo Operador",
                   "Plan (prefijo)", "SIMs", "Precio Unitario", "Moneda",
                   "Monto (P×Q)", "Comentario"]
    ws.append(encabezados)
    for col in range(1, len(encabezados) + 1):
        c = ws.cell(row=1, column=col)
        c.font = blanco_bold; c.fill = azul; c.alignment = centro

    fe_txt = resultado["fecha_ejecucion"].iloc[0].strftime("%Y-%m-%d %H:%M:%S")
    for _, r in resultado.iterrows():
        ws.append([
            r["periodo"], fe_txt, r["pais"], r["operador"], r["tipo_operador"],
            r["plan_prefijo"], r["sims_total"],
            r["precio_unitario"] if pd.notna(r["precio_unitario"]) else None,
            r["moneda"] if pd.notna(r["moneda"]) else "",
            r["monto"] if pd.notna(r["monto"]) else None,
            r["comentario"] if pd.notna(r["comentario"]) else "",
        ])

    n = len(resultado)
    primera, ultima = 2, n + 1
    for row in range(primera, ultima + 1):
        ws.cell(row=row, column=7).number_format = fmt_num
        ws.cell(row=row, column=8).number_format = fmt_dec
        ws.cell(row=row, column=10).number_format = fmt_dec

    fila_total = ultima + 1
    ws.cell(row=fila_total, column=6, value="TOTAL SIMs").font = bold
    c_tot = ws.cell(row=fila_total, column=7, value=f"=SUM(G{primera}:G{ultima})")
    c_tot.font = bold; c_tot.number_format = fmt_num
    ws.cell(row=fila_total, column=6).fill = gris
    ws.cell(row=fila_total, column=7).fill = gris

    for i, w in enumerate([9, 19, 12, 13, 14, 24, 11, 15, 9, 16, 42], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    rs = wb.create_sheet("Resumen por Moneda")
    rs["A1"] = "Reporte de Cobros por Conectividad"
    rs["A1"].font = Font(bold=True, size=14)
    rs["A2"], rs["B2"] = "Periodo:", periodo
    rs["A3"], rs["B3"] = "Fecha ejecución:", fe_txt
    rs["A4"], rs["B4"] = "Archivo origen:", (Path(archivo_origen).name if archivo_origen else "")
    for fila in (2, 3, 4):
        rs.cell(row=fila, column=1).font = bold

    hdr = 6
    for col, titulo in enumerate(["Moneda", "SIMs", "Monto Total"], start=1):
        c = rs.cell(row=hdr, column=col, value=titulo)
        c.font = blanco_bold; c.fill = azul; c.alignment = centro

    rg_mon  = f"Detalle!$I${primera}:$I${ultima}"
    rg_sims = f"Detalle!$G${primera}:$G${ultima}"
    rg_mont = f"Detalle!$J${primera}:$J${ultima}"

    monedas = sorted([m for m in resultado["moneda"].dropna().unique() if m])
    fila = hdr + 1
    for mon in monedas:
        rs.cell(row=fila, column=1, value=mon).font = normal
        s = rs.cell(row=fila, column=2, value=f'=SUMIF({rg_mon},"{mon}",{rg_sims})')
        m = rs.cell(row=fila, column=3, value=f'=SUMIF({rg_mon},"{mon}",{rg_mont})')
        s.font = normal; s.number_format = fmt_num
        m.font = normal; m.number_format = fmt_dec
        fila += 1

    rs.cell(row=fila, column=1, value="(sin precio)").font = normal
    s = rs.cell(row=fila, column=2, value=f'=SUMIF({rg_mon},"",{rg_sims})')
    s.font = normal; s.number_format = fmt_num
    fila_sinprecio = fila
    fila += 1

    rs.cell(row=fila, column=1, value="TOTAL SIMs (todas)").font = bold
    t = rs.cell(row=fila, column=2, value=f"=SUM(B{hdr+1}:B{fila_sinprecio})")
    t.font = bold; t.number_format = fmt_num
    rs.cell(row=fila, column=1).fill = gris
    rs.cell(row=fila, column=2).fill = gris

    for col, w in zip("ABC", [22, 16, 18]):
        rs.column_dimensions[col].width = w

    os.makedirs(os.path.dirname(path_salida) or ".", exist_ok=True)
    wb.save(path_salida)


# ============================================================================
# MENÚ
# ============================================================================
def _pedir_periodo():
    while True:
        anio = input("    Año (ej. 2026): ").strip()
        mes = input("    Mes (1-12): ").strip()
        if not re.fullmatch(r"20\d{2}", anio):
            print("    ⚠️ Año inválido."); continue
        if not (mes.isdigit() and 1 <= int(mes) <= 12):
            print("    ⚠️ Mes inválido."); continue
        return f"{anio}{int(mes):02d}"


def _pedir_alcance():
    print("\n  Alcance:")
    print("    1. Por periodo (año, mes)")
    print("    2. Histórico (todos los meses)")
    return input("  Opción: ").strip()


def _confirmar(msg: str) -> bool:
    return input(f"{msg} (s/n): ").strip().lower().startswith("s")


def _accion_raw(engine):
    alc = _pedir_alcance()
    if alc == "1":
        periodo = _pedir_periodo()
        archivo = localizar_archivo_periodo(periodo)
        procesar_raw(engine, archivo, periodo)
    elif alc == "2":
        pares = periodos_en_archivos()
        if not pares:
            print("  ⚠️ No hay archivos disponibles en INPUT_DIR."); return
        print(f"  Se cargarán {len(pares)} meses: {', '.join(p for p, _ in pares)}")
        if _confirmar("  ¿Continuar?"):
            for periodo, archivo in pares:
                try:
                    procesar_raw(engine, archivo, periodo)
                except Exception as e:
                    print(f"  ❌ {periodo} falló: {e}  (sigo con el resto)")
    else:
        print("  Opción inválida.")


def _accion_reporte(engine):
    alc = _pedir_alcance()
    if alc == "1":
        procesar_reporte(engine, _pedir_periodo())
    elif alc == "2":
        periodos = periodos_en_raw(engine)
        if not periodos:
            print("  ⚠️ No hay data RAW cargada. Usa primero la opción 1."); return
        print(f"  Se procesarán {len(periodos)} periodos: {', '.join(periodos)}")
        if _confirmar("  ¿Continuar?"):
            for periodo in periodos:
                try:
                    procesar_reporte(engine, periodo)
                except Exception as e:
                    print(f"  ❌ {periodo} falló: {e}  (sigo con el resto)")
    else:
        print("  Opción inválida.")


def _accion_completo(engine):
    alc = _pedir_alcance()
    if alc == "1":
        periodo = _pedir_periodo()
        archivo = localizar_archivo_periodo(periodo)
        procesar_raw(engine, archivo, periodo)
        procesar_reporte(engine, periodo)
    elif alc == "2":
        pares = periodos_en_archivos()
        if not pares:
            print("  ⚠️ No hay archivos disponibles en INPUT_DIR."); return
        print(f"  Proceso completo de {len(pares)} meses: {', '.join(p for p, _ in pares)}")
        if _confirmar("  ¿Continuar?"):
            for periodo, archivo in pares:
                try:
                    procesar_raw(engine, archivo, periodo)
                except Exception as e:
                    print(f"  ❌ RAW {periodo} falló: {e}  (sigo con el resto)")
            for periodo, _ in pares:
                try:
                    procesar_reporte(engine, periodo)
                except Exception as e:
                    print(f"  ❌ Reporte {periodo} falló: {e}  (sigo con el resto)")
    else:
        print("  Opción inválida.")


def menu():
    print("=" * 60)
    print(" 🚀 PIPELINE DE COBROS POR CONECTIVIDAD  (repo de referencia)")
    print("=" * 60)
    print("\n🔌 Conectando a PostgreSQL...")
    engine = conectar_db()
    if engine is None:
        print("\n❌ Sin conexión a la BD. No puedo continuar.")
        sys.exit(1)
    try:
        while True:
            print("\n" + "=" * 60)
            print(" MENÚ PRINCIPAL")
            print("=" * 60)
            print("  1. Cargar data RAW         (Excel → capa RAW)")
            print("  2. Generar reporte cobros  (RAW + maestro → resultado)")
            print("  3. Proceso completo        (RAW + reporte)")
            print("  4. Salir")
            op = input("\n  Opción: ").strip()
            try:
                if op == "1":
                    _accion_raw(engine)
                elif op == "2":
                    _accion_reporte(engine)
                elif op == "3":
                    _accion_completo(engine)
                elif op == "4":
                    break
                else:
                    print("  ⚠️ Opción inválida.")
            except (FileNotFoundError, ValueError) as e:
                print(f"  ❌ {e}")
            except Exception as e:
                print(f"  ❌ Error inesperado: {e}")
    finally:
        engine.dispose()
        print("\n🔌 Conexión cerrada.")


def main():
    menu()


if __name__ == "__main__":
    main()
